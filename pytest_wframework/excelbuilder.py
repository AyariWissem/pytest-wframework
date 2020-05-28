from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension


class excelBuilder:

    def __init__(self, jsonObject, path):
        self.json = jsonObject
        self.path = path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.dh = DimensionHolder(worksheet=self.ws)

    def generateExcelt(self):  # TODO rushed, needs better implementation
        """generates an excel file of the failed tests"""
        steps = ["setup", "call", "teardown"]
        report = self.json
        ws = self.ws
        dh = self.dh
        ws.cell(1, 1, "Test Case Path").font=Font(b=True)
        ws.cell(1, 2, "Test Case Name").font=Font(b=True)
        ws.cell(1, 3, "Test Case Status").font=Font(b=True)
        ws.cell(1, 4, "Messages").font=Font(b=True)
        y = {"path": 1, "name": 2, "status": 3, "Messages": 4}
        x = 2
        for test in report['tests']:
            if test['outcome'] == "failed":
                ws.cell(x, y["path"], test['nodeid'])
                ws.cell(x, y["name"], test['nodeid'].split("::")[-1])
                failedcell = ws.cell(x, y["status"], test['outcome'].upper())
                failedcell.font = Font(b=True, color="FF0000")
                messages = ""
                for step in steps:
                    if "log" in test[step]:
                        for log in test[step]["log"]:
                            print("log type:::" + str(type(log)))
                            if log["levelname"] == "MESSAGE":
                                messages += log["msg"] + "\n"
                    ws.cell(x, y["Messages"], messages)
                x += 1
        # setting the width of the columns
        for col in range(ws.min_column, ws.max_column + 1):
            dh[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=40)
        ws.column_dimensions=dh
        # saving the report
        try:
            if not (self.path[-1] == '\\'):
                self.path += '\\'
            self.wb.save(self.path + 'report.xlsx')
        except:
            print("Excel path is invalid")
            # TODO remove the json file
